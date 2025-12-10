from pathlib import Path
import csv
import time
import threading
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from typing import List, Tuple, Optional, Dict, Any
from ..utils.logger import get_logger
from ..integrations.bitbrowser import BitBrowserClient
from .automation import run_registration_flow

# 添加全局锁，用于保护CSV文件写入
_csv_locks: Dict[str, threading.Lock] = {}
 


def update_csv_status(csv_path: Path, email: str, status: str) -> bool:
    """
    更新CSV文件中指定邮箱的注册状态（第9列）
    
    参数：
        csv_path: CSV文件路径
        email: 邮箱地址
        status: 状态值（"成功" 或 "失败"）
    
    返回：
        更新成功返回True，否则返回False
    """
    log = get_logger(__name__)
    
    # 获取或创建该CSV文件的锁
    csv_key = str(csv_path.absolute())
    if csv_key not in _csv_locks:
        _csv_locks[csv_key] = threading.Lock()
    
    # 使用锁保护文件读写
    with _csv_locks[csv_key]:
        try:
            suf = csv_path.suffix.lower()
            if suf == ".csv":
                rows = []
                encs = ["utf-8-sig", "utf-8", "gb18030", "gbk", "latin-1"]
                detected = None
                try:
                    import chardet
                    raw = csv_path.read_bytes()[:4096]
                    det = chardet.detect(raw)
                    enc = det.get("encoding")
                    if enc:
                        detected = enc
                except Exception:
                    detected = None
                used = detected or encs[0]
                ok = False
                for enc in ([used] + [e for e in encs if e != used]):
                    try:
                        with csv_path.open("r", encoding=enc, newline='') as f:
                            reader = csv.reader(f)
                            rows = list(reader)
                            ok = True
                            break
                    except UnicodeDecodeError:
                        continue
                if not ok:
                    raise RuntimeError("CSV编码不受支持")
            elif suf == ".xlsx":
                import openpyxl
                wb = openpyxl.load_workbook(csv_path)
                ws = wb.worksheets[0]
                rows = []
                for r in ws.iter_rows(values_only=True):
                    rows.append(["" if v is None else str(v) for v in list(r)])
            else:
                raise RuntimeError("仅支持CSV或XLSX状态更新")
            
            if not rows:
                log.error("❌ CSV文件为空")
                return False
            
            # 确定是否有标题行
            has_header = rows[0] and "email" in rows[0][0].lower()
            start_row = 1 if has_header else 0
            
            # 查找并更新对应行
            updated = False
            for i in range(start_row, len(rows)):
                row = rows[i]
                if len(row) >= 1 and row[0].strip() == email:
                    # 扩展行到至少9列
                    while len(row) < 9:
                        row.append("")
                    # 更新第9列（索引8）
                    row[8] = status
                    rows[i] = row
                    updated = True
                    log.info(f"✅ 已更新 {email} 的状态为: {status}")
                    break
            
            if not updated:
                log.warning(f"⚠️ 未找到邮箱 {email} 对应的行")
                return False
            
            if suf == ".csv":
                with csv_path.open("w", encoding="utf-8-sig", newline='') as f:
                    writer = csv.writer(f)
                    writer.writerows(rows)
            else:
                import openpyxl
                wb = openpyxl.load_workbook(csv_path)
                ws = wb.worksheets[0]
                has_header = rows[0] and "email" in str(rows[0][0]).lower()
                start_row = 2 if has_header else 1
                for i in range(start_row, len(rows) + 1):
                    ws.cell(row=i, column=9, value=rows[i - 1][8] if len(rows[i - 1]) >= 9 else "")
                wb.save(csv_path)
            
            log.info(f"✅ CSV文件已更新: {csv_path}")
            return True
            
        except Exception as e:
            log.error(f"❌ 更新CSV状态失败: {e}")
            import traceback
            log.error(traceback.format_exc())
            return False


def load_accounts_csv(csv_path: Path, skip_success: bool = True) -> List[Dict[str, Any]]:
    """
    加载账号与代理（可选）列表
    CSV格式：邮箱账号、邮箱密码、邮箱验证码接码地址、代理ip、代理端口、代理用户名、代理密码、窗口名称、状态
    
    参数：
        csv_path: CSV文件路径
        skip_success: 是否跳过状态为"成功"的账号，默认True
    """
    log = get_logger(__name__)
    rows: List[Dict[str, Any]] = []
    suf = csv_path.suffix.lower()
    raw_rows: List[List[str]] = []
    if suf == ".csv":
        encs = ["utf-8-sig", "utf-8", "gb18030", "gbk", "latin-1"]
        detected = None
        try:
            import chardet
            raw = csv_path.read_bytes()[:4096]
            det = chardet.detect(raw)
            enc = det.get("encoding")
            if enc:
                detected = enc
        except Exception:
            detected = None
        used = detected or encs[0]
        ok = False
        for enc in ([used] + [e for e in encs if e != used]):
            try:
                with csv_path.open("r", encoding=enc) as f:
                    reader = csv.reader(f)
                    raw_rows = list(reader)
                    ok = True
                    break
            except UnicodeDecodeError:
                continue
        if not ok:
            raise RuntimeError("CSV编码不受支持")
    elif suf == ".xlsx":
        import openpyxl
        wb = openpyxl.load_workbook(csv_path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        for r in ws.iter_rows(values_only=True):
            raw_rows.append(["" if v is None else str(v) for v in list(r)])
    elif suf == ".xls":
        import xlrd
        book = xlrd.open_workbook(str(csv_path))
        sheet = book.sheet_by_index(0)
        for i in range(sheet.nrows):
            row = sheet.row_values(i)
            raw_rows.append([str(v) if v is not None else "" for v in row])
    else:
        raise RuntimeError("不支持的文件格式")
    for i, row in enumerate(raw_rows):
        
            # 跳过标题行（检查第一列是否包含'email'字样）
            if i == 0 and row and "email" in row[0].lower():
                continue
            if len(row) >= 2:
                # 检查第9列状态（索引8）
                status = row[8].strip() if len(row) >= 9 else ""
                
                if skip_success and status in ("成功", "弹窗", "弹窗弹窗"):
                    continue
                
                rec: Dict[str, Any] = {
                    "email": row[0].strip(),
                    "password": row[1].strip(),
                }
                # 第3列：邮箱验证码接码地址
                if len(row) >= 3 and row[2].strip():
                    rec["code_url"] = row[2].strip()
                # 第4-7列：代理信息（ip, port, username, password）
                if len(row) >= 5 and row[3].strip():
                    rec["host"] = row[3].strip()
                    port_str = str(row[4]).strip()
                    # 确保端口是数字
                    if port_str and port_str.isdigit():
                        rec["port"] = int(port_str)
                    else:
                        rec["port"] = None
                # 代理用户名和密码
                if len(row) >= 6 and row[5].strip():
                    rec["proxyUserName"] = row[5].strip()
                else:
                    rec["proxyUserName"] = None
                if len(row) >= 7 and row[6].strip():
                    rec["proxyPassword"] = row[6].strip()
                else:
                    rec["proxyPassword"] = None
                # 第8列（索引7）：窗口名称
                if len(row) >= 8:
                    window_name_raw = row[7]
                    if window_name_raw:
                        window_name_cleaned = ''.join(window_name_raw.split())
                        rec["windowName"] = window_name_cleaned if window_name_cleaned else None
                    else:
                        rec["windowName"] = None
                else:
                    rec["windowName"] = None
                # 第9列（索引8）：状态
                rec["status"] = status
                
                rows.append(rec)
    return rows


def mock_register(email: str, password: str) -> bool:
    # 示例：模拟注册动作，真实项目中接入目标站点的注册逻辑
    time.sleep(0.05)
    return ("@" in email) and (len(password) >= 6)


def _wait_and_switch_to_new_tab(
    client: "BitBrowserClient",
    window_id: str,
    target_url: str,
    max_retries: int = 10,
    retry_interval: float = 2.0,
    log = None
) -> bool:
    """
    等待新标签页创建完成并切换到包含目标URL的标签页
    
    参数：
        client: 比特浏览器客户端
        window_id: 窗口ID
        target_url: 目标URL（用于匹配标签页）
        max_retries: 最大重试次数
        retry_interval: 每次重试的间隔时间（秒）
        log: 日志对象
    
    返回：
        成功切换返回True，否则返回False
    """
    import re
    
    if log is None:
        log = get_logger(__name__)
    
    # 提取目标URL的域名部分用于匹配
    target_domain = re.sub(r'https?://(www\.)?', '', target_url).split('/')[0]
    log.info(f"Looking for tab with domain: {target_domain}")
    
    workspace_tab_id = None  # 保存工作台标签页ID
    
    for attempt in range(max_retries):
        try:
            log.info(f"Attempt {attempt + 1}/{max_retries} to find and switch to new tab")
            
            # 获取当前窗口的所有标签页
            tabs_response = client.get_window_tabs(window_id)
            
            if not tabs_response.get("success"):
                log.warning(f"Failed to get tabs: {tabs_response.get('msg')}")
                time.sleep(retry_interval)
                continue
            
            tabs_data = tabs_response.get("data", {})
            tabs = tabs_data.get("tabs", [])
            
            log.info(f"Found {len(tabs)} tabs in window")
            
            # 打印所有标签页的详细信息
            for i, tab in enumerate(tabs):
                tab_url = tab.get("url", "")
                tab_id = tab.get("id", "")
                tab_title = tab.get("title", "")
                is_active = tab.get("active", False)
                log.info(f"Tab {i}: ID={tab_id}, Title='{tab_title}', URL={tab_url}, Active={is_active}")
                
                # 记录工作台标签页
                if "工作台" in tab_title or "console.bitbrowser" in tab_url:
                    workspace_tab_id = tab_id
                    log.info(f"Found workspace tab: ID={tab_id}")
            
            # 查找包含目标域名的标签页
            target_tab = None
            for tab in tabs:
                tab_url = tab.get("url", "")
                tab_title = tab.get("title", "")
                
                # 检查URL或标题是否包含目标域名/关键词
                if (target_domain in tab_url or 
                    "klingai" in tab_url.lower() or 
                    "kling" in tab_title.lower()):
                    target_tab = tab
                    log.info(f"Found matching tab: Title='{tab_title}', URL={tab_url}")
                    break
            
            if target_tab:
                tab_id = target_tab.get("id")
                if tab_id:
                    # 切换到目标标签页
                    log.info(f"Switching to tab: {tab_id}")
                    switch_result = client.switch_tab(window_id, tab_id)
                    log.info(f"Switch tab result: {switch_result}")
                    
                    # 再次等待确保切换完成
                    time.sleep(1)
                    
                    # 激活窗口确保用户可见
                    try:
                        client.activate(window_id)
                        log.info("Window activated")
                    except Exception as e:
                        log.warning(f"Failed to activate window: {e}")
                    
                    log.info(f"Successfully switched to target tab!")
                    return True
            else:
                log.warning(f"Target tab not found yet, waiting...")
                time.sleep(retry_interval)
                
        except Exception as e:
            log.error(f"Error in attempt {attempt + 1}: {e}")
            time.sleep(retry_interval)
    
    # 如果所有重试失败，尝试关闭工作台标签页
    if workspace_tab_id:
        try:
            log.info(f"Failed to switch tab after {max_retries} attempts")
            log.info(f"Trying fallback: closing workspace tab (ID={workspace_tab_id})...")
            
            close_result = client.close_tab(window_id, workspace_tab_id)
            log.info(f"Close workspace tab result: {close_result}")
            
            # 等待标签页关闭
            time.sleep(2)
            
            # 激活窗口，Kling AI标签应该自动显示
            try:
                client.activate(window_id)
                log.info("Window activated after closing workspace tab")
            except Exception as e:
                log.warning(f"Failed to activate window: {e}")
            
            log.info("Workspace tab closed, Kling AI tab should now be visible")
            return True
            
        except Exception as e:
            log.error(f"Failed to close workspace tab: {e}")
    
    log.error(f"Failed to switch to new tab after all attempts")
    return False


def register_accounts_batch(
    csv_path: Path,
    runtime_dir: Path,
    concurrency: int = 3,
    interval_ms: int = 300,
    bitbrowser_base_url: Optional[str] = None,
    platform_url: Optional[str] = None,
    bitbrowser_password: Optional[str] = None,
    auto_xpaths: Optional[Dict[str, str]] = None,
    dry_run: bool = False,
    browser_mode: str = "bitbrowser",
    stop_flag: Optional[Dict[str, bool]] = None,
) -> str:
    log = get_logger(__name__)
    rows = load_accounts_csv(csv_path)

    # Playwright模式：直接执行自动化，加入多轮重试
    if browser_mode == "playwright":
        log.info(f"🎭 Playwright模式 - 将使用本地浏览器 + 随机指纹")
        ok, fail = 0, 0
        outputs: List[str] = []
        max_rounds = 5
        for round_idx in range(1, max_rounds + 1):
            if stop_flag and stop_flag.get("stop", False):
                log.info("⏹️ 用户中断，停止所有轮次")
                break
            rows = load_accounts_csv(csv_path, skip_success=True)
            if not rows:
                log.info("🎉 所有账号均已成功，无需继续重试")
                break
            log.info(f"🔁 开始第 {round_idx}/{max_rounds} 轮注册，待处理 {len(rows)} 个账号")
            for idx, rec in enumerate(rows, 1):
                if stop_flag and stop_flag.get("stop", False):
                    log.info("⚠️ 检测到中断信号，停止注册")
                    break
                email = rec.get("email")
                password = rec.get("password")
                code_url = rec.get("code_url")
                host = rec.get("host")
                port = rec.get("port")
                puser = rec.get("proxyUserName")
                ppass = rec.get("proxyPassword")
                log.info(f"\n{'='*60}")
                log.info(f"Processing account {idx}/{len(rows)}: {email}")
                log.info(f"{'='*60}")
                try:
                    auto_ok = run_registration_flow(
                        email=email,
                        password=password,
                        runtime_dir=runtime_dir,
                        xpaths=auto_xpaths or {},
                        proxy={
                            "host": host,
                            "port": port,
                            "username": puser,
                            "password": ppass,
                        },
                        platform_url=platform_url or "https://klingai.com",
                        code_url=code_url,
                        attach_ws=None,
                        dry_run=dry_run,
                        browser_mode="playwright",
                    )
                    if auto_ok:
                        ok += 1
                        success_msg = f"SUCCESS {idx}/{len(rows)}: {email}"
                        outputs.append(success_msg)
                        log.info(f"✅ {success_msg}")
                        update_csv_status(csv_path, email, "成功")
                    else:
                        fail += 1
                        fail_msg = f"FAIL {idx}/{len(rows)}: {email} - automation failed"
                        outputs.append(fail_msg)
                        log.error(f"❌ {fail_msg}")
                        update_csv_status(csv_path, email, "失败")
                except Exception as e:
                    fail += 1
                    outputs.append(f"ERROR {idx}/{len(rows)}: {email}: {e}")
                    log.error(f"ERROR processing {email}: {e}")
                    update_csv_status(csv_path, email, "失败")
                time.sleep(interval_ms / 1000.0)
        final_rows = load_accounts_csv(csv_path, skip_success=False)
        total_accounts = len(final_rows)
        final_success = sum(1 for r in final_rows if (r.get('status') == '成功'))
        final_failed = total_accounts - final_success
        summary = (
            f"\n{'='*60}\nBatch Registration Complete\n{'='*60}\n"
            f"Final Success: {final_success} | Final Failed: {final_failed} | Total: {total_accounts}\n"
            f"Attempts -> Success: {ok} | Failed: {fail}\n"
            f"{'='*60}"
        )
        log.info(summary)
        return "\n".join(outputs + [summary])

    # 比特浏览器模式：需要比特浏览器客户端
    if not bitbrowser_base_url:
        # 如果没有比特浏览器URL且不是Playwright模式，使用mock注册
        ok, fail = 0, 0
        outputs: List[str] = []
        with ThreadPoolExecutor(max_workers=max(1, concurrency)) as ex:
            futures = {}
            for rec in rows:
                email = rec.get("email")
                password = rec.get("password")
                fut = ex.submit(mock_register, email, password)
                futures[fut] = email
                time.sleep(interval_ms / 1000.0)
            for fut in as_completed(futures):
                email = futures[fut]
                try:
                    if fut.result():
                        ok += 1
                        outputs.append(f"SUCCESS {email}")
                    else:
                        fail += 1
                        outputs.append(f"FAIL {email}")
                except Exception as e:
                    fail += 1
                    outputs.append(f"ERROR {email}: {e}")
        summary = f"done. success={ok}, fail={fail}"
        log.info(summary)
        return "\n".join(outputs + [summary])

    client = BitBrowserClient(bitbrowser_base_url)
    ok, fail = 0, 0
    outputs: List[str] = []
    
    log.info(f"开始批量注册 {len(rows)} 个账号，并发数: {concurrency}")
    
    # 🔴 使用线程池实现并发执行
    def force_cleanup_window(client: BitBrowserClient, window_id: str, password: Optional[str], logger) -> bool:
        """
        强制清理窗口：确保窗口被关闭和删除，即使遇到网络错误也要多次尝试
        
        Args:
            client: BitBrowserClient实例
            window_id: 窗口ID
            password: 比特浏览器密码
            logger: 日志记录器
        
        Returns:
            bool: 是否成功清理
        """
        import time
        
        try:
            # 最多尝试5次清理
            for attempt in range(5):
                try:
                    logger.info(f"🔄 清理窗口尝试 {attempt + 1}/5: {window_id}")
                    
                    # 1. 尝试关闭窗口
                    try:
                        close_result = client.close_window(window_id)
                        if close_result.get("success"):
                            logger.info(f"✅ 窗口已关闭: {window_id}")
                        else:
                            logger.warning(f"⚠️ 关闭窗口失败: {close_result.get('msg', '未知错误')}")
                    except Exception as close_err:
                        logger.warning(f"⚠️ 关闭窗口异常: {close_err}")
                    
                    # 2. 等待确保窗口关闭
                    time.sleep(2)
                    
                    # 3. 尝试删除窗口
                    try:
                        delete_result = client.delete_window(window_id, password)
                        if delete_result.get("success"):
                            logger.info(f"✅ 窗口已删除: {window_id}")
                            return True  # 成功删除
                        else:
                            logger.warning(f"⚠️ 删除窗口失败: {delete_result.get('msg', '未知错误')}")
                    except Exception as del_err:
                        logger.warning(f"⚠️ 删除窗口异常: {del_err}")
                    
                    # 4. 如果不是最后一次尝试，等待后重试
                    if attempt < 4:
                        wait_time = 2 ** attempt  # 1s, 2s, 4s, 8s
                        logger.info(f"⏳ 等待 {wait_time} 秒后重试...")
                        time.sleep(wait_time)
                        
                except Exception as attempt_err:
                    logger.warning(f"⚠️ 清理尝试 {attempt + 1} 异常: {attempt_err}")
                    if attempt < 4:
                        time.sleep(2 ** attempt)
            
            logger.error(f"❌ 窗口清理最终失败: {window_id}")
            return False
            
        except Exception as e:
            logger.error(f"❌ 窗口清理过程发生严重异常: {e}")
            return False
    
    
    def process_account(idx: int, rec: Dict[str, Any]) -> Tuple[bool, str, str, Optional[str]]:
        """
        处理单个账号注册
        返回: (success, email, message, window_id)
        """
        # 检查中断标志（在开始前检查）
        if stop_flag and stop_flag.get("stop", False):
            email = rec.get("email", "")
            log.info(f"⏹️ 检测到中断信号，跳过账号处理: {email}")
            return (False, email, f"SKIP: 用户中断 {idx}/{len(rows)}: {email}", None)
        
        email = rec.get("email")
        password = rec.get("password")
        code_url = rec.get("code_url")
        window_name = rec.get("windowName")
        window_id = None
        
        log.info(f"\n{'='*60}")
        log.info(f"处理账号 {idx}/{len(rows)}: {email}")
        if window_name:
            log.info(f"📁 窗口名称: {window_name}")
        else:
            log.info("📁 窗口名称: 未设置")
        log.info(f"{'='*60}")
        
        try:
            payload: Dict[str, Any] = {
                "userName": email,
                "password": password,
                "url": platform_url or "https://klingai.com",
                "proxyMethod": 2,
                "browserFingerPrint": {},
            }
            
            if window_name:
                payload["remark"] = window_name
                payload["name"] = window_name
                payload["windowName"] = window_name
                payload["browserAlias"] = window_name
            
            # 设置代理
            host = rec.get("host")
            port = rec.get("port")
            puser = rec.get("proxyUserName")
            ppass = rec.get("proxyPassword")
            if host and port:
                payload.update({
                    "proxyType": "socks5",
                    "host": host,
                    "port": int(port),
                })
                if puser:
                    payload["proxyUserName"] = puser
                if ppass:
                    payload["proxyPassword"] = ppass
            
            # 检查中断标志（在创建窗口前检查）
            if stop_flag and stop_flag.get("stop", False):
                log.info(f"⏹️ 检测到中断信号，取消窗口创建: {email}")
                return (False, email, f"SKIP: 用户中断 {idx}/{len(rows)}: {email}", None)
            
            window_id = None
            try:
                r = client.create_window(payload)
                if not r.get("success"):
                    data = r.get("data", {})
                    temp_window_id = data.get("id") or ""
                    return (False, email, f"FAIL: {r.get('msg')}", temp_window_id if temp_window_id else None)
                data = r.get("data", {})
                window_id = data.get("id") or ""
                if not window_id:
                    return (False, email, "FAIL: No window ID", None)
                log.info(f"🆕 创建新窗口: {window_id}")
            except Exception as acquire_err:
                return (False, email, f"ERROR {email}: {acquire_err}", None)
            
            # 检查中断标志（在打开窗口前检查）
            if stop_flag and stop_flag.get("stop", False):
                log.info(f"⏹️ 检测到中断信号，关闭并删除已创建的窗口: {email} (ID: {window_id})")
                # 立即清理已创建的窗口
                force_cleanup_window(client, window_id, bitbrowser_password, log)
                return (False, email, f"SKIP: 用户中断 {idx}/{len(rows)}: {email}", None)
            
            # 打开窗口
            open_result = client.open_window(window_id)
            log.info(f"Window opened for {email}: {open_result}")
            time.sleep(5)
            
            # 激活窗口确保显示在前台
            try:
                activate_result = client.activate(window_id)
                log.info(f"✅ 窗口已激活: {activate_result}")
            except Exception as activate_err:
                log.warning(f"⚠️ 窗口激活失败: {activate_err}")
                # 即使激活失败也继续执行，避免中断整个流程
            
            # 等待窗口稳定
            time.sleep(2)
            
            # 获取WebSocket
            ws = None
            if open_result.get("success"):
                ws_data = open_result.get("data", {})
                ws = ws_data.get("ws") or ws_data.get("webSocketDebuggerUrl")
                if ws:
                    log.info(f"✅ Got WebSocket: {ws}")
                else:
                    if "http" in ws_data.get("http", ""):
                        http_url = ws_data.get("http")
                        import re
                        port_match = re.search(r':(\d+)', http_url)
                        if port_match:
                            port = port_match.group(1)
                            ws = f"ws://127.0.0.1:{port}/devtools/browser"
                            log.info(f"✅ Constructed WebSocket: {ws}")
            
            # 再次激活窗口确保显示在前台
            try:
                client.activate(window_id)
                log.info("✅ 窗口已再次激活")
            except Exception as activate_err:
                log.warning(f"⚠️ 窗口再次激活失败: {activate_err}")
                # 即使激活失败也继续执行，避免中断整个流程
            
            # 检查中断标志（在执行自动化前检查）
            if stop_flag and stop_flag.get("stop", False):
                log.info(f"⏹️ 检测到中断信号，关闭并删除已打开的窗口: {email} (ID: {window_id})")
                # 立即清理已打开的窗口
                force_cleanup_window(client, window_id, bitbrowser_password, log)
                return (False, email, f"SKIP: 用户中断 {idx}/{len(rows)}: {email}", None)
            
            # 执行自动化注册
            auto_ok = run_registration_flow(
                email=email,
                password=password,
                runtime_dir=runtime_dir,
                xpaths=auto_xpaths or {},
                proxy={
                    "host": host,
                    "port": port,
                    "username": puser,
                    "password": ppass,
                },
                platform_url=platform_url or "https://klingai.com",
                code_url=code_url,
                attach_ws=ws,
                dry_run=dry_run,
                browser_mode=browser_mode,
            )
            
            if auto_ok:
                if window_id:
                    try:
                        close_result = client.close_window(window_id)
                        log.info(f"close_window: {close_result}")
                    except Exception:
                        pass
                return (True, email, f"SUCCESS {idx}/{len(rows)}: {email}", window_id)
            else:
                if window_id:
                    try:
                        force_cleanup_window(client, window_id, bitbrowser_password, log)
                    except Exception:
                        pass
                return (False, email, f"FAIL {idx}/{len(rows)}: {email} - automation failed", window_id)
                
        except Exception as e:
            # 异常时可删除未成功的窗口
            log.error(f"❌ 注册过程发生异常: {e}")
            if window_id:
                try:
                    # 弹窗场景：关闭并删除
                    if str(e) == "POPUP_DETECTED":
                        force_cleanup_window(client, window_id, bitbrowser_password, log)
                        return (False, email, f"POPUP {idx}/{len(rows)}: {email}", window_id)
                except Exception:
                    pass
                force_cleanup_window(client, window_id, bitbrowser_password, log)
            return (False, email, f"ERROR {email}: {e}", None)
    
    # 使用线程池并发执行
    # 多轮重试：最多5轮，直到全部成功或中断
    max_rounds = 5
    for round_idx in range(1, max_rounds + 1):
        if stop_flag and stop_flag.get("stop", False):
            log.info("⏹️ 用户中断，停止所有轮次")
            break
        rows = load_accounts_csv(csv_path, skip_success=True)
        if not rows:
            log.info("🎉 所有账号均已成功，无需继续重试")
            break
        log.info(f"🔁 开始第 {round_idx}/{max_rounds} 轮注册，待处理 {len(rows)} 个账号")
        with ThreadPoolExecutor(max_workers=max(1, concurrency)) as executor:
            futures = {}
            for idx, rec in enumerate(rows, 1):
                if stop_flag and stop_flag.get("stop", False):
                    log.info("⚠️ 检测到中断信号，停止提交新任务")
                    break
                future = executor.submit(process_account, idx, rec)
                futures[future] = rec.get("email")
                time.sleep(interval_ms / 1000.0)
            for future in as_completed(futures):
                if stop_flag and stop_flag.get("stop", False):
                    log.info("⚠️ 检测到中断信号，取消剩余任务")
                    for f in futures:
                        if not f.done():
                            f.cancel()
                    break
                email = futures[future]
                try:
                    success, result_email, message, window_id = future.result()
                    outputs.append(message)
                    if success:
                        ok += 1
                        log.info(f"✅ {message}")
                        update_csv_status(csv_path, result_email, "成功")
                    else:
                        fail += 1
                        log.error(f"❌ {message}")
                        if message.startswith("POPUP "):
                            update_csv_status(csv_path, result_email, "弹窗弹窗")
                        else:
                            update_csv_status(csv_path, result_email, "失败")
                except Exception as e:
                    fail += 1
                    error_msg = f"ERROR {email}: {e}"
                    outputs.append(error_msg)
                    log.error(error_msg)
                    update_csv_status(csv_path, email, "失败")
    
    # 最终统计
    final_rows = load_accounts_csv(csv_path, skip_success=False)
    total_accounts = len(final_rows)
    final_success = sum(1 for r in final_rows if (r.get('status') == '成功'))
    final_failed = total_accounts - final_success
    summary = (
        f"\n{'='*60}\nBatch Registration Complete\n{'='*60}\n"
        f"Final Success: {final_success} | Final Failed: {final_failed} | Total: {total_accounts}\n"
        f"Attempts -> Success: {ok} | Failed: {fail}\n"
        f"{'='*60}"
    )
    log.info(summary)
    try:
        import os as _os
        if _os.environ.get("KL_CLEANUP_GUARD", "0") == "1" and bitbrowser_base_url:
            client = BitBrowserClient(bitbrowser_base_url)
            cleanup_residual_windows(client, csv_path, log)
    except Exception:
        pass
    return "\n".join(outputs + [summary])
def cleanup_residual_windows(client: BitBrowserClient, csv_path: Path, logger) -> None:
    try:
        rows = load_accounts_csv(csv_path, skip_success=False)
        names = set([r.get("windowName") for r in rows if r.get("status") == "失败" and r.get("windowName")])
        if not names:
            return
        resp = client.list_windows()
        if not resp.get("success"):
            return
        data = resp.get("data", {})
        lst = data.get("list") or data.get("data") or []
        for w in lst:
            wname = (w.get("remark") or w.get("name") or w.get("windowName") or "").replace(" ", "")
            if wname and wname in names:
                wid = w.get("id") or w.get("windowId") or ""
                if wid:
                    try:
                        client.close_window(wid)
                    except Exception:
                        pass
                    try:
                        client.delete_window(wid)
                        logger.info(f"🧹 清理残留窗口: {wid} ({wname})")
                    except Exception:
                        pass
    except Exception:
        pass
